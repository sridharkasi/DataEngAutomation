import openpyxl
from pyspark.sql import SparkSession
from pyspark.sql.types import StructType, StructField, StringType, IntegerType
# from pyspark.sql.functions import iloc
import allure
from allure_commons._core import plugin_manager
from allure_pytest.listener import AllureListener

global executionDict
global spark
global srcdict
global trgdict
global SrcCount
global TrgCount
global sourceDf
global targetDf

# def test_main():
def executecase(Action, SourceFormat, SourceFilePath, TargetFilePath, TargetFormat, SQL, TestcaseName ):
  ret = ""
  if (SourceFilePath!=None) and (TargetFilePath!=None):
        spark=CreateSparkSession()
        if(Action!="Schema"):
            df = loadsparkdata(spark, SourceFormat, SourceFilePath)
            sourceDf = df
  else:
        print("--Source or target File path not provided for the test case. Pls Check--"+ TestcaseName)
        return
  if (TargetFilePath != None):
      spark = CreateSparkSession()
      df = loadsparkdata(spark, TargetFormat, TargetFilePath)
      targetDf = df
  else:
      print("TargetFilePath Blank for the test case" + TestcaseName)
  if(Action == 'RowCount'):
        RowcountTest(sourceDf, targetDf, TestcaseName, SQL, spark)
  elif(Action == 'ColCount'):
        ColcountTest(sourceDf, targetDf, TestcaseName,SQL, spark)
  elif (Action == 'DupCount'):
      print("")
      DuplicateCount(sourceDf, targetDf, TestcaseName, SQL, spark)
  elif (Action == 'Schema'):
      print("")
      ret= ComapreSchema(targetDf, SourceFilePath, TestcaseName )
  elif (Action == 'SchemaDF'):
      SchemaDifference(targetDf, sourceDf)
  elif (Action == 'CompareTable'):
      ComapareDatatable_EachRow(targetDf, sourceDf, SQL, spark, TestcaseName)
  return ret
# create spark Session for the tests
def CreateSparkSession():
    spark = SparkSession.builder \
        .appName("Automation") \
        .master("local[1]") \
        .config("spark.jars", "./jars/spark-xml_2.12-0.9.0.jar") \
        .config("spark.executor.extraClassPath", "./jars/spark-xml_2.12-0.9.0.jar") \
        .config("spark.executor.extraLibrary", "./jars/spark-xml_2.12-0.9.0.jar") \
        .config("spark.driver.extraClassPath", "./jars/spark-xml_2.12-0.9.0.jar") \
        .getOrCreate()
    return spark
# Load spark data for the tests
def loadsparkdata(spark, Format,filepath):
    schema = StructType([
        StructField('PolNumber', StringType(), True),
        StructField('LineOfBusiness', StringType(), True)
    ])

    if(Format=='XML'):
        df = spark.read.format("com.databricks.spark.xml") \
            .option("rowTag", "Policy").load(filepath, schema=schema)
        return df
    elif(Format=='csv'):
        df = spark \
            .read \
            .option("header", "true") \
            .option("inferSchema", "true") \
            .csv(filepath)
        return df
@allure.step
def RowcountTest(sourceDf, targetDf, TestcaseName,SQL, spark):
    print("--------------RFL Insurance TB Vs CFL Insurance-----------------")
    print("Testcase: [" + TestcaseName + "]")
    if (SQL == None):
        print("------------Source Table-----RFL Insurance------------------------------------")
        print("Record Count: "+ sourceDf.count())
        print("------------Target Table-----CFL Insurance----------------------------------------")
        print("Record Count: "+ targetDf.count())
        srccou = sourceDf.count()
        trgcou = targetDf.count()
    else:
        sourceDf.createOrReplaceTempView("table")
        sqldt = spark.sql(SQL)
        # sqldt.show()
        srccou = sqldt.first()['count(1)']

        targetDf.createOrReplaceTempView("table")
        sqlTarg = spark.sql(SQL)
        # sqlTarg.show()
        trgcou = sqlTarg.first()['count(1)']

    if (srccou == trgcou):
        sourceDf.select(sourceDf.columns[0:5]).show(10)
        # dif.select(dif.columns[0:5]).show()
        print("Source Vs Target Count: Pass :[" + str(sourceDf.count()) + "]Vs[" + str(targetDf.count())+"]")
        print("Source Vs Target Count: Pass :[" + str(srccou) + "]Vs[" + str(trgcou) + "]")
        print("-----------------------------------------------------------------------")
        # targetDf.show()
    else:
        # targetDf.show()
        targetDf.select(sourceDf.columns[0:5]).show(10)
        # print("Source Vs Target Count: Fail " + str(sourceDf.count()) + "Vs" + str(targetDf.count())
        print("Source Vs Target Count: Fail :[" + str(srccou) + "]Vs[" + str(trgcou) + "]")
@allure.step
def ColcountTest(sourceDf, targetDf, TestcaseName,SQL, spark):
    print("---------------RBenefits Table---------------------------------------")
    print("Testcase: [" + TestcaseName + "]")
    print("Source Data Columns: ["+str (sourceDf.columns)+"]")
    print("Target Data Columns: [" +str(targetDf.columns) + "]")
    if(len(sourceDf.columns)==len(targetDf.columns)):
        print("Source Data Columns : Pass: ["+str (len(sourceDf.columns))+"] Vs Target Data Columns: [" +str(len(targetDf.columns)) + "] Matching")
    else:
        print("Source Data Columns: Fail: [" + str(sourceDf.columns.count()) + "] Vs Target Data Columns: [" + str(targetDf.columns.count()) + "] Not Matching > --Fail")
@allure.step
def DuplicateCount(sourceDf, targetDf, TestcaseName,SQL, spark):
    print("----------------RBenefits Table--------------------------------------")
    print("Testcase: [" + TestcaseName + "]")
    print("------------------------------------------------------")
    if (SQL==None):
        print("No sql is Provided")
        return
    else:
        sourceDf.createOrReplaceTempView("table")
        SrcDup = spark.sql(SQL)
        SrcDupcount=SrcDup.count()
        print("---SourceDB Duplicate Records Found :" + str(SrcDupcount))
        SrcDup.show()
        targetDf.createOrReplaceTempView("table")
        TrgDup = spark.sql(SQL)
        TrgDupcount = TrgDup.count()
        # print("---TargetDB Duplicate Records Found :" + str(TrgDupcount) )
        # TrgDup.show()
@allure.step
def ComapreSchema(targetDf, SourceFilePath, TestcaseName ):
    print("---------------RBenefits Table-----------------------------------")
    print("Testcase: [" + TestcaseName + "]")
    print("------------------------------------------------------")
    targetDf.printSchema()
    srcdict = dict()
    trgdict = dict()
    dictionary1= getsourceschema(srcdict, SourceFilePath, TestcaseName)
    dictionary2= gettargetschema(targetDf, trgdict)
    print("---Compare Source Vs Target Schema-------------------------------------")
    # compareschema(srcdict, trgdict)
    returnval = compareschema(dictionary1, dictionary2)
    return returnval

def gettargetschema(df, trgdict):
        for item in df.schema.fields:
            x = str(item)
            y = x.split("(")
            z = y[1].split(",")
            # print(z[0])
            key = z[0]
            value = z[1]
            trgdict[str(key)] = [value]
        return trgdict
def getsourceschema(srcdict,SourceFilePath,TestcaseName):
    filename = SourceFilePath
    wb = openpyxl.load_workbook(filename)
    ws = wb[TestcaseName]
    rw = ws.max_row
    for i in range(1, rw + 1):
        Keyname = ws.cell(row=i, column=1)
        Keyvalue = ws.cell(row=i, column=2)
        srcdict[Keyname.value] = Keyvalue.value
    return srcdict
@allure.step
def compareschema(srcdict, trgdict):
    x = len(srcdict)
    y = len(trgdict)

    for key in srcdict.keys():
        # print(key)
        if key in trgdict.keys():
            # str1 = srcdict[key]
            str1 = str(srcdict[key]).split("(")
            # print (str1[0])
            # if(trgdict[key]==srcdict[key]):
            if(str(trgdict[key]).find(str1[0])):
                # print("Data Type matches")
                print(key+ " " +str(trgdict[key]) + "Vs [" + str(srcdict[key])+ "] matches")
                print ("---------------------------------------------------------------")
            else:
                # print("Data Type not matches")
                print(key+ " " +str(trgdict[key]) + "Vs [" + str(srcdict[key])+ "] Not matches")
                print ("-------------------------------------------------------------------")
    if (x == y):
        print("Length of the schema in both tables matching > Pass")
        print(len(srcdict))
        print(len(trgdict))
        return True
    else:
        print("Length of the schema in both tables not matching > Fail")
        print(len(srcdict))
        print(len(trgdict))
        return False
@allure.step
def SchemaDifference(targetDf, sourceDf):
    df_schema = sourceDf.schema.fields
    data_schema = targetDf.schema.fields
    df_names = [x.name.lower() for x in df_schema]
    data_names = [x.name.lower() for x in data_schema]
    if (df_schema != data_schema):
        col_diff = set(df_names) ^ set(data_names)
        print (col_diff)

        # col_list = [(x[0].name, x[0].dataType) for x in map(df_schema, data_schema) if((x[0] is not None and x[0].name.lower() in col_diff) or x[1].name.lower() in col_diff)]
        # for i in col_list:
        #     if i[0] in df_names:
        #         data = data.withColumn("%s" % i[0],_lit_doc(None).cast(i[1]))
        #     else:
        #         df = df.withColumn("%s" % i[0],_lit_doc(None).cast(i[1]))

    else:
        print("Nothing to do")

@allure.step
def ComapareDatatable_EachRow(targetDf, sourceDf, SQL, spark, TestcaseName):
    print("-------------------TestCase: "+ TestcaseName + "---------RFL Insurance----------")
    if(SQL==None):
        print(sourceDf.exceptAll(targetDf))
        # print(targetDf.exceptAll(sourceDf))
        # targetDf.exceptAll(sourceDf).show(1)
        if(int(sourceDf.exceptAll(targetDf).count()) > 0):
            print("Below Record found not matching with source data-- Fail")
            dif= sourceDf.exceptAll(targetDf)
            dif.select(dif.columns[0:5]).show()
        else:
            print("--All the Records from target table matching with source data-- pass")
    else:
        sourceDf.createOrReplaceTempView("table1")
        targetDf.createOrReplaceTempView("table")
        Dif= spark.sql(SQL)
        if (int(Dif.count()) > 0):
            print("--Below Record found not matching with source data--> Fail")
            Dif.select(Dif.columns[0:5]).show()
            # Dif.show()
        else:
            print("All the Records from target table matching with source data-->pass")
    # x = 0
    # flag=0
    # # print(sdf.head(2)[0])
    # RowCount = targetDf.count()
    # for i in sourceDf.collect():
    #     print(x)
    #     row = targetDf.head(RowCount)[x]
    #     if (i == row):
    #         print("row matches")
    #     else:
    #         print("Table row not matches at id:" + str(x))
    #         flag=1
    #     x = x + 1
    # if (flag==0):
    #     print("All the cell data values are matching in both the table")


    # filename = SourceFilePath
    # wb = openpyxl.load_workbook(filename)
    # ws = wb[TestcaseName]
    # rw = ws.max_row
    # max_column = ws.max_column
    #     # executionDict = dict()
    # for i in range(1, rw + 1):
    #     Keyname = ws.cell(row=i, column=1)
    #     Keyvalue = ws.cell(row=i, column=2)
    #     srcdict[Keyname.value] = Keyvalue.value
    # x=0
    # for j in targetDf.schema:
    #     print(j[0])

    # sourceDf.printSchema()

#     if (i==1):
#         Format = SourceFormat
#         Path=SourceFilePath
#         if(Action!='Schema'):
#             df = loadsparkdata(spark,Format, SourceFilePath)
#     else:
#         Format = TargetFormat
#         Path = TargetFilePath
#         if (Action != 'Schema'):
#             df = loadsparkdata(spark, Format, TargetFilePath)
#     if (Action == 'RowCount'):
#        print("-----" + TestcaseName + "----------"+ Action + "----")
#        function1(df, TestcaseName)
#        # df.show()
#        # print(df.count())
#     elif(Action == 'ColCount'):
#        print("-----" + TestcaseName + "----------" + Action + "----")
#        function2(df, TestcaseName)
#        df.printSchema()
#     elif (Action == 'DupCount'):
#        print("-----" + TestcaseName + "----------" + Action + "----")
#        function3(df, TestcaseName)
#     elif (Action == 'Schema'):
#        print("-----" + TestcaseName + "----------" + Action + "----")
#        function4(spark, Format, SourceFilePath, TargetFilePath,TestcaseName)
#     else:
#      print("specify action")
#
#     if (Action != 'Schema'):
#         if (i==1):
#             sdf = df
#             # print("-------source Data-----------")
#             print("-------Source table------------")
#             if(SQL==None):
#                 sdf.show()
#                 srccnt= sdf.count()
#             else:
#                 sdf.createOrReplaceTempView("table")
#                 sqldt= spark.sql(SQL)
#                 sqldt.show()
#                 srccnt = sqldt.count()
#             # sourcecount = sdf.count()
#             sourcecount = srccnt
#
#         else:
#             tdf = df
#             print("-------Target table------------")
#             if (SQL == None):
#                 tdf.show()
#                 trgcnt=tdf.count()
#             else:
#                 tdf.createOrReplaceTempView("table")
#                 trg = spark.sql(SQL)
#                 trg.show()
#                 trgcnt = trg.count()
#             targetcount = trgcnt
#     else:
#         print("")
#     spark.stop()
#   if (Action != 'Schema'):
#     assert (targetcount==sourcecount)
#
# @allure.step
# def function1(df,TestcaseName):
#     print ("-----"+TestcaseName+"----------")
#     df.show(10)
#     # allure.attach(df.show(10))
#     # allure.dynamic.description(df.show(10))
#     # allure.description(df.show(10))
#     # allure.title(df.show(10))
#     return df
# @allure.step
# def function2(df,TestcaseName):
#     df.show(10)
#     # allure.step("--->Data Frame --->")
#     # allure.description(df.show(10))
#     # allure.attach(df.show(10))
#     return df
# @allure.step
# def function3(df, TestcaseName):
#     print("-----" + TestcaseName + "----------")
#     df.show(10)
#     return df
# @allure.step
# def function4(spark, Format, SourceFilePath, TargetFilePath,TestcaseName):
#     print("-----" + TestcaseName + "----------")
#     srcdict = dict()
#     trgdict = dict()
#     df = loadsparkdata(spark, Format, TargetFilePath)
#     getsourceschema(srcdict,SourceFilePath,TestcaseName)
#     gettargetschema(df, trgdict)
#     compareschema(srcdict, trgdict)
#     # sourceschema(SourceFilePath)
#     # targetschema(TargetFilePath)
#
# @allure.step
# def createsparksession():
#     spark = SparkSession.builder \
#         .appName("Automation") \
#         .master("master") \
#         .config("spark.jars", "file:///C://spark3//jars//spark-xml_2.12-0.9.0.jar") \
#         .config("spark.executor.extraClassPath", "file:///C://spark3//jars//spark-xml_2.12-0.9.0.jar") \
#         .config("spark.executor.extraLibrary", "file:///C://spark3//jars//spark-xml_2.12-0.9.0.jar") \
#         .config("spark.driver.extraClassPath", "file:///C://spark3//jars//spark-xml_2.12-0.9.0.jar") \
#         .getOrCreate()
#     # return spark
# @allure.step
# def loadsparkdata(spark, Format,filepath):
#     schema = StructType([
#         StructField('PolNumber', StringType(), True),
#         StructField('LineOfBusiness', StringType(), True)
#     ])
#
#     if(Format=='XML'):
#         df = spark.read.format("com.databricks.spark.xml") \
#             .option("rowTag", "Policy").load(filepath, schema=schema)
#         return df
#     elif(Format=='csv'):
#         df = spark \
#             .read \
#             .option("header", "true") \
#             .option("inferSchema", "true") \
#             .csv(filepath)
#         return df
# # def sourceschema(SourceFilePath):
#
# # def targetschema(TargetFilePath):
# @allure.step
# def gettargetschema(df, trgdict):
#     for item in df.schema.fields:
#          # print(item)
#          x = str (item)
#          y = x.split("(")
#          # print(y[1])
#          z = y[1].split(",")
#          print(z[0])
#          key = z[0]
#          value =z[1]
#          # print(z[1])
#          # print("***********")
#          trgdict[str(key)]= [value]
# @allure.step
# def getsourceschema(srcdict,SourceFilePath,TestcaseName):
#     filename = SourceFilePath
#     wb = openpyxl.load_workbook(filename)
#     ws = wb[TestcaseName]
#     rw = ws.max_row
#     max_column = ws.max_column
#     # executionDict = dict()
#     for i in range(1, rw + 1):
#         Keyname = ws.cell(row=i, column=1)
#         Keyvalue = ws.cell(row=i, column=2)
#         srcdict[Keyname.value] = Keyvalue.value
# @allure.step
# def compareschema(srcdict, trgdict):
#     x = len(srcdict)
#     y = len(trgdict)
#     print(len(srcdict))
#     print(len(trgdict))
#     print("Length of the schema in both tables")
#     # assert (x == y)
#     # diff = set(srcdict) - set(trgdict)
#     # print("Differences Found: "+str(diff))
#     for key in srcdict.keys():
#         print(key)
#         if key in trgdict.keys():
#             if(trgdict[key]==srcdict[key]):
#                 print("Data Type matches")
#                 print(str(trgdict[key]) + "Vs" + str(srcdict[key]))
#                 print ("---------------------")
#             else:
#                 print("Data Type not matches")
#                 print(str(trgdict[key]) + "Vs" + str(srcdict[key]))
#                 print("---------------------")
#     assert (x == y)
    # for key, value in srcdict.iteritems():
    #     for dict1_value in dict1_values:
    #         try:
    #             dict2_value = trgdict[dict1_value]
    #             print
    #             str(dict1_key) + '\t' + str(dict1_value) + '\t' + str(dict2_value)
    #         except KeyError:
    #             pass












