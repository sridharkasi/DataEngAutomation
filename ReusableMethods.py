import openpyxl
from pyspark.sql import SparkSession
from pyspark.sql.types import StructType, StructField, StringType, IntegerType
import allure
from allure_commons._core import plugin_manager
from allure_pytest.listener import AllureListener
global executionDict
global spark
global srcdict
global trgdict
global SrcCount
global TrgCount

# def test_main():
def executecase(Action, SourceFormat, SourceFilePath, TargetFilePath, TargetFormat, SQL, TestcaseName ):

  for i in (1,2):
    spark = SparkSession.builder \
        .appName("Automation") \
        .master("local[1]") \
        .config("spark.jars", "./jars/spark-xml_2.12-0.9.0.jar") \
        .config("spark.executor.extraClassPath", "./jars/spark-xml_2.12-0.9.0.jar") \
        .config("spark.executor.extraLibrary", "./jars/spark-xml_2.12-0.9.0.jar") \
        .config("spark.driver.extraClassPath", "./jars/spark-xml_2.12-0.9.0.jar") \
        .getOrCreate()
    sqlq = SQL
    if (i==1):
        Format = SourceFormat
        Path=SourceFilePath
        if(Action!='Schema'):
            df = loadsparkdata(spark,Format, SourceFilePath)
    else:
        Format = TargetFormat
        Path = TargetFilePath
        if (Action != 'Schema'):
            df = loadsparkdata(spark, Format, TargetFilePath)
    if (Action == 'RowCount'):
       print("-----" + TestcaseName + "----------"+ Action + "----")
       function1(df, TestcaseName)
       # df.show()
       # print(df.count())
    elif(Action == 'ColCount'):
       print("-----" + TestcaseName + "----------" + Action + "----")
       function2(df, TestcaseName)
       df.printSchema()
    elif (Action == 'DupCount'):
       print("-----" + TestcaseName + "----------" + Action + "----")
       function3(df, TestcaseName)
    elif (Action == 'Schema'):
       print("-----" + TestcaseName + "----------" + Action + "----")
       function4(spark, Format, SourceFilePath, TargetFilePath,TestcaseName)
    else:
     print("specify action")

    if (Action != 'Schema'):
        if (i==1):
            sdf = df
            # print("-------source Data-----------")
            print("-------Source table------------")
            if(SQL==None):
                sdf.show()
                srccnt= sdf.count()
            else:
                sdf.createOrReplaceTempView("table")
                sqldt= spark.sql(SQL)
                sqldt.show()
                srccnt = sqldt.count()
            # sourcecount = sdf.count()
            sourcecount = srccnt

        else:
            tdf = df
            print("-------Target table------------")
            if (SQL == None):
                tdf.show()
                trgcnt=tdf.count()
            else:
                tdf.createOrReplaceTempView("table")
                trg = spark.sql(SQL)
                trg.show()
                trgcnt = trg.count()
            targetcount = trgcnt
    else:
        print("")
    spark.stop()
  if (Action != 'Schema'):
    assert (targetcount==sourcecount)

@allure.step
def function1(df,TestcaseName):
    print ("-----"+TestcaseName+"----------")
    df.show(10)
    # allure.attach(df.show(10))
    # allure.dynamic.description(df.show(10))
    # allure.description(df.show(10))
    # allure.title(df.show(10))
    return df
@allure.step
def function2(df,TestcaseName):
    df.show(10)
    # allure.step("--->Data Frame --->")
    # allure.description(df.show(10))
    # allure.attach(df.show(10))
    return df
@allure.step
def function3(df, TestcaseName):
    print("-----" + TestcaseName + "----------")
    df.show(10)
    return df
@allure.step
def function4(spark, Format, SourceFilePath, TargetFilePath,TestcaseName):
    print("-----" + TestcaseName + "----------")
    srcdict = dict()
    trgdict = dict()
    df = loadsparkdata(spark, Format, TargetFilePath)
    getsourceschema(srcdict,SourceFilePath,TestcaseName)
    gettargetschema(df, trgdict)
    compareschema(srcdict, trgdict)
    # sourceschema(SourceFilePath)
    # targetschema(TargetFilePath)

@allure.step
def createsparksession():
    spark = SparkSession.builder \
        .appName("Automation") \
        .master("master") \
        .config("spark.jars", "file:///C://spark3//jars//spark-xml_2.12-0.9.0.jar") \
        .config("spark.executor.extraClassPath", "file:///C://spark3//jars//spark-xml_2.12-0.9.0.jar") \
        .config("spark.executor.extraLibrary", "file:///C://spark3//jars//spark-xml_2.12-0.9.0.jar") \
        .config("spark.driver.extraClassPath", "file:///C://spark3//jars//spark-xml_2.12-0.9.0.jar") \
        .getOrCreate()
    # return spark
@allure.step
def loadsparkdata(spark, Format,filepath):
    schema = StructType([
        StructField('PolNumber', StringType(), True),
        StructField('LineOfBusiness', StringType(), True)
    ])

    if(Format=='XML'):
        df = spark \
        # .read \
        # .format('com.databricks.spark.xml') \
        # .options(rootTag=rootTag) \
        # .options(rowTag=rowTag) \
        # .options(rowTag=rowTag1) \
        # .load('SampleXmlData.xml')
        df = spark.read.format("com.databricks.spark.xml") \
            .option("rowTag", "Policy").load(filepath, schema=schema)
        return df
    elif(Format=='csv'):
        df = spark \
            .read \
            .option("header", "true") \
            .option("inferSchema", "true") \
            .csv(filepath)
        return df
# def sourceschema(SourceFilePath):

# def targetschema(TargetFilePath):
@allure.step
def gettargetschema(df, trgdict):
    for item in df.schema.fields:
         # print(item)
         x = str (item)
         y = x.split("(")
         # print(y[1])
         z = y[1].split(",")
         print(z[0])
         key = z[0]
         value =z[1]
         # print(z[1])
         # print("***********")
         trgdict[str(key)]= [value]
@allure.step
def getsourceschema(srcdict,SourceFilePath,TestcaseName):
    filename = SourceFilePath
    wb = openpyxl.load_workbook(filename)
    ws = wb[TestcaseName]
    rw = ws.max_row
    max_column = ws.max_column
    # executionDict = dict()
    for i in range(1, rw + 1):
        Keyname = ws.cell(row=i, column=1)
        Keyvalue = ws.cell(row=i, column=2)
        srcdict[Keyname.value] = Keyvalue.value
@allure.step
def compareschema(srcdict, trgdict):
    x = len(srcdict)
    y = len(trgdict)
    print(len(srcdict))
    print(len(trgdict))
    print("Length of the schema in both tables")
    # assert (x == y)
    # diff = set(srcdict) - set(trgdict)
    # print("Differences Found: "+str(diff))
    for key in srcdict.keys():
        print(key)
        if key in trgdict.keys():
            if(trgdict[key]==srcdict[key]):
                print("Data Type matches")
                print(str(trgdict[key]) + "Vs" + str(srcdict[key]))
                print ("---------------------")
            else:
                print("Data Type not matches")
                print(str(trgdict[key]) + "Vs" + str(srcdict[key]))
                print("---------------------")
    assert (x == y)
    # for key, value in srcdict.iteritems():
    #     for dict1_value in dict1_values:
    #         try:
    #             dict2_value = trgdict[dict1_value]
    #             print
    #             str(dict1_key) + '\t' + str(dict1_value) + '\t' + str(dict2_value)
    #         except KeyError:
    #             pass












