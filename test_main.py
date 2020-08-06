import openpyxl
from pyspark.sql import SparkSession
# import ReusableMethods
import ReusableFunctions
from pyspark.sql.types import StructType, StructField, StringType, IntegerType
global executionDict
global spark
import allure
# def test_main():
@allure.parent_suite("Data Engg Automation Suite")
@allure.suite("Main test Suite")
def test_ReadExecXl():
    # global spark
    filename="./Data/Execution.xlsx"
    wb = openpyxl.load_workbook(filename)
    ws = wb['Sheet1']
    rw = ws.max_row
    max_column=ws.max_column
    executionDict = dict()
    passflag = True
    for i in range(2, rw + 1):
        returnflag = True
        for j in range(1, max_column+1 ):
            flag = 0
            execution = ws.cell(row=i, column=1)
            # print(execution.value)
            if (execution.value == 'Y'):
                Keyname = ws.cell(row=1, column=j)
                keyvalue= ws.cell(row=i, column=j)
                executionDict[Keyname.value]= keyvalue.value
                # odict['Policy' + str(i)] = i['PolNumber']
            # executecase(executionDict["Action"], executionDict["SourceFormat"])
            # executecase(executionDict["Action"], executionDict["SourceFormat"], executionDict["SourceFilePath"],executionDict["TargetFilePath"], executionDict["TargetFormat"], executionDict["SQL"])
            else:
                flag=1
                break

        if(flag==0):
           returnflag=ReusableFunctions.executecase(executionDict["Action"], executionDict["SourceFormat"], executionDict["SourceFilePath"], executionDict["TargetFilePath"], executionDict["TargetFormat"], executionDict["SQL"], executionDict["TestcaseName"])
        if(returnflag==True or returnflag==None):
            print("")
        else:
            passflag=False
    assert passflag







